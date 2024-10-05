#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import math
import openpyxl
import numpy as np
import datetime
import pandas as pd


months_to_num = {
    "Jan": "01",
    "Feb": "02",
    "Mar": "03",
    "Apr": "04",
    "May": "05",
    "Jun": "06",
    "Jul": "07",
    "Aug": "08",
    "Sep": "09",
    "Oct": "10",
    "Nov": "11",
    "Dec": "12",
}


class Song:
    """
    A class to represent a song with various attributes and methods to format and process these attributes.

    Attributes:
    -----------
    songID : str
        A unique identifier for the song derived from its name.
    rank : str
        The rank of the song, formatted appropriately.
    htmlID : str
        A unique identifier for the song based on its live date and media type.
    length : str
        The length of the song, formatted appropriately.
    length_DT : datetime.timedelta
        The length of the song as a timedelta object.
    media : str
        The media type of the song (e.g., YouTube, Live, Twitch).
    date_str : str
        The date string derived from the htmlID and media type.
    date_DT : datetime.datetime
        The date of the song as a datetime object.
    date_YM_DT : datetime.datetime
        The year and month of the song's date as a datetime object.
    name : str
        The name of the song, formatted appropriately.
    tempo : str
        The tempo of the song, formatted appropriately.
    comment : str
        The comment associated with the song, formatted appropriately.
    choral : str
        The choral information of the song, formatted appropriately.
    genre : str
        The genre of the song.
    URL : str
        The URL of the song.
    when_ranked_DT : datetime.datetime
        The date when the song was ranked as a datetime object.
    when_ranked_YM_DT : datetime.datetime
        The year and month when the song was ranked as a datetime object.
    live_title : str
        The live title of the song.
    live_comment : str
        The live comment associated with the song, formatted appropriately.
    """

    def __init__(
        self,
        name,
        media,
        live_date,
        rank,
        genre,
        tempo,
        length,
        comment,
        choral,
        URL,
        when_ranked,
        live_title,
        live_comment,
    ):
        self.songID = self.get_songID(name)
        self.rank = self.format_rank(rank)
        self.htmlID = self.get_htmlID(live_date, media)
        self.length = self.format_length(length, self.songID)
        self.length_DT = self.get_length_DT(self.length)
        self.media = media
        self.date_str = self.get_date_str(self.htmlID, media)
        self.date_DT = datetime.datetime(
            year=int("20" + self.date_str[0:2]),
            month=int(self.date_str[2:4]),
            day=int(self.date_str[4:6]),
        )
        self.date_YM_DT = datetime.datetime(year=self.date_DT.year, month=self.date_DT.month, day=1)
        self.name = self.format_name(name)
        self.tempo = self.format_tempo(tempo)
        self.comment = self.format_comment(comment)
        self.choral = self.format_choral(choral)
        self.genre = genre
        self.URL = URL
        self.when_ranked_DT = datetime.datetime(
            year=int("20" + str(when_ranked)[0:2]),
            month=int(str(when_ranked)[2:4]),
            day=int(str(when_ranked)[4:6]),
        )
        self.when_ranked_YM_DT = datetime.datetime(
            year=self.when_ranked_DT.year, month=self.when_ranked_DT.month, day=1
        )
        self.live_title = live_title
        self.live_comment = self.format_live_comment(live_comment)

    def get_htmlID(self, live_date, media):
        """
        Generates a unique htmlID based on the live date and media type.

        Parameters:
        live_date (str): The live date of the song.
        media (str): The media type of the song (e.g., YouTube, Live, Twitch).

        Returns:
        str: The generated htmlID.
        """
        if media in ["YouTube", "Live"]:
            day = f"0{live_date[4]}" if "," in live_date[4:6] else live_date[4:6]
            if media == "YouTube":
                return f"s{live_date[-2:]}{months_to_num[live_date[:3]]}{day}"
            else:
                return f"s{live_date[-2:]}{months_to_num[live_date[:3]]}{day}l"
        elif media == "Twitch":
            return f"s{live_date}"

    def get_date_str(self, htmlID, media):
        """
        Extracts the date string from the htmlID based on the media type.

        Parameters:
        htmlID (str): The htmlID of the song.
        media (str): The media type of the song (e.g., YouTube, Live, Twitch).

        Returns:
        str: The extracted date string.
        """
        if media == "YouTube":
            return htmlID[1:]
        elif media in ["Live", "Twitch"]:
            return htmlID[1:-1]

    def get_length_DT(self, length):
        """
        Converts the length of the song from a string to a timedelta object.

        Parameters:
        length (str): The length of the song as a string.

        Returns:
        datetime.timedelta: The length of the song as a timedelta object.
        """
        if (length not in [np.nan]) and len(length) > 3:
            if length[1] == "'":
                return datetime.timedelta(minutes=int(length[0]), seconds=int(length[2] + length[3]))
            elif length[2] == "'":
                return datetime.timedelta(
                    minutes=int(length[0] + length[1]),
                    seconds=int(length[3] + length[4]),
                )
            else:
                return datetime.timedelta(minutes=0, seconds=0)
        else:
            return datetime.timedelta(minutes=0, seconds=0)

    def format_tempo(self, tempo):
        """
        Formats the tempo of the song.

        Parameters:
        tempo (str): The tempo of the song.

        Returns:
        str: The formatted tempo.
        """
        if "into" in tempo:
            tempo = tempo.split(" into ")[0]

        if tempo in ["medium", "medium "]:
            return "Med"
        elif tempo in ["medium fast", "slow fast"]:
            return "Fmed"
        elif tempo in ["medium slow"]:
            return "Smed"
        elif tempo in [
            "fast",
            "fast af",
            "fast fast",
            "ultra fast",
            "ULTRA FAST",
            "jsp",
        ]:
            return "Fast"
        elif tempo in [
            "slow",
            "slow ",
            "sloooow",
            "hors du temps (14'30 en vrai)",
            "out of time (14'30 in reality)",
            "-",
        ]:
            return "Slow"
        else:
            print(tempo, "not defined")
            sys.exit()

    def format_comment(self, comment):
        """
        Formats the comment associated with the song.

        Parameters:
        comment (str): The comment associated with the song.

        Returns:
        str: The formatted comment.
        """
        for c in ["\n", '"']:
            comment = comment.replace(c, "")
        comment.replace("…", "...")
        return comment

    def format_choral(self, choral):
        """
        Formats the choral information of the song.

        Parameters:
        choral (str): The choral information of the song.

        Returns:
        str: The formatted choral information.
        """
        try:
            math.isnan(choral)
            choral = "-"
        except:
            pass
        choral = choral.replace("\n", "")
        return choral

    def format_name(self, name):
        """
        Formats the name of the song.

        Parameters:
        name (str): The name of the song.

        Returns:
        str: The formatted name.
        """
        name = name.replace('"', "")
        return name

    def get_songID(self, name):
        """
        Generates a unique songID from the name of the song.

        Parameters:
        name (str): The name of the song.

        Returns:
        str: The generated songID.
        """
        for c in [
            "-",
            " ",
            "  ",
            "'",
            "(",
            ")",
            ".",
            '"',
            ",",
            "&",
            ":",
            "!",
            "?",
            "\\",
            "/",
        ]:
            name = name.replace(c, "")
        if name[0].isdigit():
            name = "n" + name
        return name

    def format_rank(self, rank):
        """
        Formats the rank of the song.

        Parameters:
        rank (str): The rank of the song.

        Returns:
        str: The formatted rank.
        """
        rank = rank.replace("-", "I")
        if rank == "S I D":
            rank = "D"
        return rank

    def format_length(self, length, songID):
        """
        Formats the length of the song based on the songID.

        Parameters:
        length (str): The length of the song as a string.
        songID (str): The unique identifier of the song.

        Returns:
        str: The formatted length.
        """
        if songID == "Dramaticevent":
            length = "14'30"
        return length

    def format_live_comment(self, live_comment):
        """
        Formats the live comment associated with the song.

        Parameters:
        live_comment (str): The live comment associated with the song.

        Returns:
        str: The formatted live comment.
        """
        if live_comment not in [np.nan]:
            return live_comment
        return "-"


def get_df_from_xls(xls_file, media, EN=False):
    """
    Extracts data from an Excel file and returns it as a pandas DataFrame.

    Parameters:
    xls_file (str): The path to the Excel file.
    media (str): The media type of the songs (e.g., YouTube, Live, Twitch).
    EN (bool): A flag to determine which sheet to read from the workbook.
               If False, reads from "Sheet1" (French). If True, reads from "Sheet2" (English).

    Returns:
    pd.DataFrame: A DataFrame containing the extracted song data.
    """
    song_list = []
    date_set = False

    # Load the workbook and select the appropriate sheet
    wrkbk = openpyxl.load_workbook(xls_file)
    if not EN:
        sh = wrkbk["Sheet1"]
    else:
        sh = wrkbk["Sheet2"]

    line_start = 5

    # Iterate through the rows of the sheet
    for n_line, row in enumerate(
        sh.iter_rows(min_row=line_start, min_col=1, max_row=1500, max_col=10),
        line_start,
    ):
        line = []
        for cell in row:
            if cell.value is None:
                line.append(np.nan)
            else:
                line.append(cell.value)
        if row[2].value is not None:
            song_URL = row[2].hyperlink.target

        title_or_date = line[0]
        song_name = line[2]
        if type(song_name) == int:
            song_name = str(song_name)
        song_rank = line[3]
        song_genre = line[4]
        song_tempo = line[5]
        song_length = line[6]
        song_comment = line[7]
        song_choral = line[8]

        # Skip rows where song_name is NaN
        try:
            math.isnan(song_name)
            continue
        except:
            # Check if the row contains a date or a title
            try:
                math.isnan(title_or_date)
            except:
                if date_set:
                    date_set = False
                else:
                    live_title = line[0]
                    live_comment = line[1]
                    song_date = sh.cell(n_line + 1, 1).value
                    song_when_ranked = line[9]
                    if math.isnan(song_when_ranked):
                        song_when_ranked = 990101
                    date_set = True

        # Handle choral information
        try:
            math.isnan(song_choral)
        except:
            if '"' in song_choral:
                song_choral = song_choral.replace('"', sh.cell(n_line - 1, 9).value)

        # Append the song information to the list
        song_list.append(
            Song(
                song_name,
                media,
                song_date,
                song_rank,
                song_genre,
                song_tempo,
                song_length,
                song_comment,
                song_choral,
                song_URL,
                song_when_ranked,
                live_title,
                live_comment,
            )
        )

    # Convert the list of Song objects to a DataFrame
    return pd.DataFrame([vars(v) for v in song_list])


def get_unique_songID(df):
    """
    Ensures that each songID in the DataFrame is unique by appending a count to duplicate songIDs.

    Parameters:
    df (pd.DataFrame): The DataFrame containing song data with a "songID" column.

    Returns:
    pd.DataFrame: The DataFrame with unique songIDs.
    """
    print("\nMultiple songIDs format:")
    songIDs = []
    for index, row in df.iterrows():
        songID = row["songID"]
        songIDs.append(songID)
        if songIDs.count(songID) > 1:
            print(f"Multiple {songID} ({songIDs.count(songID)})")
            df.at[index, "songID"] = songID + str(songIDs.count(songID))
    print("Done")
    return df


def print_simple_stats(df):
    """
    Prints simple statistics about the songs in the DataFrame.

    Parameters:
    df (pd.DataFrame): The DataFrame containing song data.

    Prints:
    Various statistics including the count of songs by rank, number of live streams, total music length,
    average music time per stream, longest music time for a stream, average song length, longest song length,
    average number of songs per stream, highest number of songs for a stream, and estimated live streams and songs left to rank.
    """
    print("\n_______ Simple stats _______")

    # Print song counts by rank
    ranks = ["S", "A+", "A", "B+", "B", "C+", "C", "D", "I"]
    print(f"\nAny rank : ({len(df)})")
    for rank in ranks:
        print(f'{rank:>2} : ({len(df.query(f"rank == \'{rank}\'"))})')

    # Print live stream statistics
    num_streams = len(df.groupby("htmlID"))
    total_length = df["length_DT"].sum()
    avg_length_per_stream = total_length / num_streams
    longest_stream_length = df[["length_DT", "htmlID"]].groupby("htmlID").sum().max().iloc[0]
    longest_stream_title = df[["length_DT", "live_title"]].groupby("live_title").sum().idxmax().iloc[0]
    longest_stream_htmlID = df[["length_DT", "htmlID"]].groupby("htmlID").sum().idxmax().iloc[0]

    print(f"\nNumber of Live Streams done: {num_streams} streams")
    print(f"Total music ranked: {len(df)} songs")
    print(f"Total music length ranked: {str(total_length)}\n")
    print(f"Average music time per streams: {str(avg_length_per_stream)[7:15]}")
    print(f"Longest music time for a stream: {str(longest_stream_length)[7:15]}, {longest_stream_title}")
    if longest_stream_htmlID != "s201209":
        print("NEW RECORD!\n")
    else:
        print()

    # Print song length statistics
    avg_song_length = total_length / len(df)
    longest_song_length = df["length_DT"].max()
    longest_song_name = df.loc[df["length_DT"].idxmax()]["name"]

    print(f"Average song length: {str(avg_song_length)[7:15]}")
    print(f"Longest song length: {str(longest_song_length)[7:15]} with {longest_song_name}\n")

    # Print song count per stream statistics
    avg_songs_per_stream = len(df) / num_streams
    max_songs_per_stream = df[["length_DT", "live_title"]].groupby("live_title").size().max()
    max_songs_stream_title = df[["length_DT", "live_title"]].groupby("live_title").size().idxmax()

    print(f"Average number of songs per stream: {avg_songs_per_stream:.2f} songs")
    print(f"Highest number of songs for a stream: {max_songs_per_stream} songs, {max_songs_stream_title}")
    if longest_stream_htmlID != "s180211":
        print("NEW RECORD!\n")
    else:
        print()

    # Print estimated remaining statistics
    estimated_live_left = 108 - num_streams
    estimated_songs_left = estimated_live_left * avg_songs_per_stream
    estimated_length_left = estimated_live_left * avg_length_per_stream

    print(f"Number of Live Stream left to rank (estimation): {estimated_live_left}")
    print(f"Number of songs left to rank (estimation): {estimated_songs_left:.0f}")
    print(f"Total music length left to rank (estimation): {str(estimated_length_left)[7:15]}\n")
